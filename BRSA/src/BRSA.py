"""
# ==============================================================================
# Author:       Victor S Caricatte De Araújo
# Email:        victorleniwys@gmail.com or victorsc@ufmg.br
# Intitution:   Universidade federal de Minas Gerais
# Version:      1.0.8
# Date:         Jun, 3
# ...................................
# ==============================================================================
"""

import sys
import os
import pandas as pd
import numpy as np
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                             QTabWidget, QLabel, QPushButton, QFileDialog, QTextEdit, 
                             QComboBox, QProgressBar, QMessageBox, QCheckBox, QSpinBox,
                             QGroupBox, QRadioButton, QFileDialog, QDialog, QDialogButtonBox)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import seaborn as sns
from bioinfokit.analys import stat, get_data
import gprofiler
import mplcursors
from statsmodels.stats.multitest import multipletests
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import tempfile
import shutil
import json
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
import traceback
from matplotlib.widgets import Cursor
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar

class RNAseqAnalysisWorker(QThread):
    progress_updated = pyqtSignal(int)
    analysis_completed = pyqtSignal(dict)
    error_occurred = pyqtSignal(str)

    def __init__(self, counts_file, metadata_file, conditions, min_counts=10, pval_cutoff=0.05):
        super().__init__()
        self.counts_file = counts_file
        self.metadata_file = metadata_file
        self.conditions = conditions
        self.min_counts = min_counts
        self.pval_cutoff = pval_cutoff
        self.results = {}

    def run(self):
        try:
            # Step 1: Load and validate data
            self.progress_updated.emit(10)
            counts_df = pd.read_csv(self.counts_file, index_col=0)
            metadata = pd.read_csv(self.metadata_file)
            
            # Validate metadata structure
            if 'condition' not in metadata.columns:
                raise ValueError("Metadata file must contain a 'condition' column")
            
            # Validate sample matching
            counts_samples = set(counts_df.columns)
            metadata_samples = set(metadata['sample'])
            
            if not counts_samples.issuperset(metadata_samples):
                missing_samples = metadata_samples - counts_samples
                raise ValueError(f"Counts file is missing samples found in metadata: {missing_samples}")
            
            if not metadata_samples.issuperset(counts_samples):
                extra_samples = counts_samples - metadata_samples
                print(f"Warning: Metadata is missing samples found in counts: {extra_samples}")
            
            # Filter low count genes
            counts_df = counts_df.loc[counts_df.sum(axis=1) >= self.min_counts]
            
            # Store raw counts
            self.results['raw_counts'] = counts_df
            self.results['metadata'] = metadata
            
            # Step 2: Normalize data (CPM)
            self.progress_updated.emit(30)
            norm_counts = counts_df.div(counts_df.sum(axis=0), axis=1) * 1e6
            self.results['normalized_counts'] = norm_counts
            
            # Step 3: Differential expression analysis
            self.progress_updated.emit(50)
            res = stat()
            res.df = norm_counts.T
            res.df['group'] = metadata['condition']
            res.t_test('group', self.conditions[0], self.conditions[1])
            
            de_results = res.result
            de_results['gene'] = norm_counts.index
            de_results = de_results.set_index('gene')
            
            # Add log2 fold change
            de_results['log2fc'] = np.log2(de_results['mean_1'] / de_results['mean_2'])
            
            # Apply multiple testing correction
            rejected, pvals_corrected, _, _ = multipletests(
                de_results['p-value'], 
                alpha=self.pval_cutoff, 
                method='fdr_bh'
            )
            de_results['p-adj'] = pvals_corrected
            de_results['significant'] = rejected
            
            self.results['de_results'] = de_results
            
            # Step 4: PCA
            self.progress_updated.emit(70)
            from sklearn.decomposition import PCA
            from sklearn.preprocessing import StandardScaler
            
            # Log transform normalized counts
            X = np.log2(norm_counts.T + 1)
            X = StandardScaler().fit_transform(X)
            
            pca = PCA(n_components=2)
            principal_components = pca.fit_transform(X)
            self.results['pca'] = {
                'components': principal_components,
                'explained_variance': pca.explained_variance_ratio_
            }
            
            # Step 5: Prepare enrichment input
            self.progress_updated.emit(90)
            sig_genes = de_results[de_results['significant']].index.tolist()
            self.results['significant_genes'] = sig_genes
            
            self.analysis_completed.emit(self.results)
            self.progress_updated.emit(100)
            
        except FileNotFoundError as e:
            self.error_occurred.emit(f"File not found: {e.filename}")
        except pd.errors.EmptyDataError:
            self.error_occurred.emit("The CSV file is empty or corrupted")
        except ValueError as e:
            self.error_occurred.emit(f"Data validation error: {str(e)}")
        except KeyError as e:
            self.error_occurred.emit(f"Missing required column: {str(e)}")
        except Exception as e:
            self.error_occurred.emit(f"Unexpected error: {str(e)}\n{traceback.format_exc()}")

class SaveSessionDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Save Session")
        
        layout = QVBoxLayout()
        
        self.label = QLabel("Select session file location:")
        layout.addWidget(self.label)
        
        self.file_layout = QHBoxLayout()
        self.file_edit = QTextEdit()
        self.file_edit.setMaximumHeight(30)
        self.file_edit.setReadOnly(True)
        self.browse_btn = QPushButton("Browse...")
        
        self.file_layout.addWidget(self.file_edit)
        self.file_layout.addWidget(self.browse_btn)
        layout.addLayout(self.file_layout)
        
        self.options_group = QGroupBox("Save Options")
        options_layout = QVBoxLayout()
        
        self.include_data_check = QCheckBox("Include raw data in session")
        self.include_data_check.setChecked(True)
        options_layout.addWidget(self.include_data_check)
        
        self.options_group.setLayout(options_layout)
        layout.addWidget(self.options_group)
        
        self.button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)
        layout.addWidget(self.button_box)
        
        self.setLayout(layout)
        
        self.browse_btn.clicked.connect(self.browse_file)
        
    def browse_file(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "Save Session", "", "JSON Files (*.json)")
        if file_path:
            self.file_edit.setPlainText(file_path)
    
    def get_file_path(self):
        return self.file_edit.toPlainText()
    
    def include_raw_data(self):
        return self.include_data_check.isChecked()

class RNAseqAnalysisPlatform(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("BRSA - Basic RNA-Seq Analysis")
        self.setGeometry(100, 100, 1200, 800)
        
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        
        self.layout = QVBoxLayout()
        self.central_widget.setLayout(self.layout)
        
        self.create_menu()
        self.create_main_interface()
        self.create_help_tab()
        
        self.worker = None
        self.analysis_results = None
        self.temp_dir = tempfile.mkdtemp()
        self.current_session_file = None
        
    def create_menu(self):
        menubar = self.menuBar()
        
        # File menu
        file_menu = menubar.addMenu("File")
        
        open_counts_action = file_menu.addAction("Open Counts File")
        open_counts_action.triggered.connect(self.open_counts_file)
        
        open_metadata_action = file_menu.addAction("Open Metadata File")
        open_metadata_action.triggered.connect(self.open_metadata_file)
        
        file_menu.addSeparator()
        
        save_session_action = file_menu.addAction("Save Session")
        save_session_action.triggered.connect(self.save_session)
        
        load_session_action = file_menu.addAction("Load Session")
        load_session_action.triggered.connect(self.load_session)
        
        file_menu.addSeparator()
        
        export_menu = file_menu.addMenu("Export")
        
        export_report_action = export_menu.addAction("Export PDF Report")
        export_report_action.triggered.connect(self.export_report)
        
        export_excel_action = export_menu.addAction("Export Excel Workbook")
        export_excel_action.triggered.connect(self.export_excel)
        
        exit_action = file_menu.addAction("Exit")
        exit_action.triggered.connect(self.close)
        
        # Analysis menu
        analysis_menu = menubar.addMenu("Analysis")
        
        run_analysis_action = analysis_menu.addAction("Run Analysis")
        run_analysis_action.triggered.connect(self.run_analysis)
        
    def create_main_interface(self):
        self.tabs = QTabWidget()
        self.layout.addWidget(self.tabs)
        
        # Input Tab
        self.input_tab = QWidget()
        self.input_layout = QVBoxLayout()
        self.input_tab.setLayout(self.input_layout)
        
        # File selection
        file_group = QGroupBox("Input Files")
        file_layout = QVBoxLayout()
        
        self.counts_file_label = QLabel("No counts file selected")
        file_layout.addWidget(self.counts_file_label)
        
        self.browse_counts_btn = QPushButton("Browse Counts File")
        self.browse_counts_btn.clicked.connect(self.open_counts_file)
        file_layout.addWidget(self.browse_counts_btn)
        
        self.metadata_file_label = QLabel("No metadata file selected")
        file_layout.addWidget(self.metadata_file_label)
        
        self.browse_metadata_btn = QPushButton("Browse Metadata File")
        self.browse_metadata_btn.clicked.connect(self.open_metadata_file)
        file_layout.addWidget(self.browse_metadata_btn)
        
        file_group.setLayout(file_layout)
        self.input_layout.addWidget(file_group)
        
        # Analysis parameters
        param_group = QGroupBox("Analysis Parameters")
        self.param_layout = QVBoxLayout()
        
        condition_layout = QHBoxLayout()
        self.condition1_combo = QComboBox()
        self.condition1_combo.setPlaceholderText("Select condition 1")
        condition_layout.addWidget(QLabel("Condition 1:"))
        condition_layout.addWidget(self.condition1_combo)
        condition_layout.addStretch()
        
        self.condition2_combo = QComboBox()
        self.condition2_combo.setPlaceholderText("Select condition 2")
        condition_layout.addWidget(QLabel("Condition 2:"))
        condition_layout.addWidget(self.condition2_combo)
        self.param_layout.addLayout(condition_layout)
        
        filter_layout = QHBoxLayout()
        self.min_counts_spin = QSpinBox()
        self.min_counts_spin.setMinimum(0)
        self.min_counts_spin.setMaximum(1000)
        self.min_counts_spin.setValue(10)
        filter_layout.addWidget(QLabel("Min counts per gene:"))
        filter_layout.addWidget(self.min_counts_spin)
        filter_layout.addStretch()
        
        self.pval_cutoff_spin = QSpinBox()
        self.pval_cutoff_spin.setMinimum(0)
        self.pval_cutoff_spin.setMaximum(100)
        self.pval_cutoff_spin.setValue(5)
        filter_layout.addWidget(QLabel("P-value cutoff (%):"))
        filter_layout.addWidget(self.pval_cutoff_spin)
        self.param_layout.addLayout(filter_layout)
        
        param_group.setLayout(self.param_layout)
        self.input_layout.addWidget(param_group)
        
        # Run button
        self.run_btn = QPushButton("Run Analysis")
        self.run_btn.clicked.connect(self.run_analysis)
        self.input_layout.addWidget(self.run_btn)
        
        # Progress bar
        self.progress_bar = QProgressBar()
        self.input_layout.addWidget(self.progress_bar)
        
        # Results tabs
        self.results_tab = QWidget()
        self.results_layout = QVBoxLayout()
        self.results_tab.setLayout(self.results_layout)
        
        self.results_tabs = QTabWidget()
        self.results_layout.addWidget(self.results_tabs)
        
        # DE Results tab
        self.de_tab = QWidget()
        self.de_layout = QVBoxLayout()
        self.de_tab.setLayout(self.de_layout)
        
        self.de_text = QTextEdit()
        self.de_text.setReadOnly(True)
        self.de_layout.addWidget(self.de_text)
        
        # Visualization tab
        self.viz_tab = QWidget()
        self.viz_layout = QVBoxLayout()
        self.viz_tab.setLayout(self.viz_layout)
        
        self.viz_canvas = FigureCanvas(Figure(figsize=(10, 8)))
        self.viz_layout.addWidget(self.viz_canvas)
        
        # Add navigation toolbar for zoom/pan
        self.toolbar = NavigationToolbar(self.viz_canvas, self)
        self.viz_layout.addWidget(self.toolbar)
        
        viz_controls = QHBoxLayout()
        self.viz_type_combo = QComboBox()
        self.viz_type_combo.addItems(["PCA Plot", "Heatmap", "Volcano Plot"])
        self.viz_type_combo.currentTextChanged.connect(self.update_visualization)
        viz_controls.addWidget(self.viz_type_combo)
        
        self.save_plot_btn = QPushButton("Save Plot")
        self.save_plot_btn.clicked.connect(self.save_current_plot)
        viz_controls.addWidget(self.save_plot_btn)
        
        self.viz_layout.addLayout(viz_controls)
        
        # Enrichment tab
        self.enrichment_tab = QWidget()
        self.enrichment_layout = QVBoxLayout()
        self.enrichment_tab.setLayout(self.enrichment_layout)
        
        # Enrichment options
        enrichment_options = QGroupBox("Enrichment Options")
        options_layout = QVBoxLayout()
        
        self.source_group = QGroupBox("Data Sources")
        source_layout = QVBoxLayout()
        
        self.go_bp_check = QCheckBox("GO Biological Process")
        self.go_bp_check.setChecked(True)
        source_layout.addWidget(self.go_bp_check)
        
        self.go_mf_check = QCheckBox("GO Molecular Function")
        source_layout.addWidget(self.go_mf_check)
        
        self.go_cc_check = QCheckBox("GO Cellular Component")
        source_layout.addWidget(self.go_cc_check)
        
        self.kegg_check = QCheckBox("KEGG Pathways")
        source_layout.addWidget(self.kegg_check)
        
        self.reactome_check = QCheckBox("Reactome Pathways")
        source_layout.addWidget(self.reactome_check)
        
        self.do_check = QCheckBox("Disease Ontology")
        source_layout.addWidget(self.do_check)
        
        self.source_group.setLayout(source_layout)
        options_layout.addWidget(self.source_group)
        
        enrichment_options.setLayout(options_layout)
        self.enrichment_layout.addWidget(enrichment_options)
        
        # Enrichment results
        self.enrichment_text = QTextEdit()
        self.enrichment_text.setReadOnly(True)
        self.enrichment_layout.addWidget(self.enrichment_text)
        
        self.run_enrichment_btn = QPushButton("Run Enrichment Analysis")
        self.run_enrichment_btn.clicked.connect(self.run_enrichment_analysis)
        self.enrichment_layout.addWidget(self.run_enrichment_btn)
        
        # Add tabs to main interface
        self.tabs.addTab(self.input_tab, "Input")
        self.tabs.addTab(self.results_tab, "Results")
        
    def create_help_tab(self):
        self.help_tab = QWidget()
        self.help_layout = QVBoxLayout()
        self.help_tab.setLayout(self.help_layout)
        
        help_text = """
        <h1>BRSA (Basic RNA-Seq Analysis) Help</h1>
        
        <h2>Getting Started</h2>
        <p>1. Select your counts file (CSV format with genes as rows and samples as columns)</p>
        <p>2. Select your metadata file (CSV format with sample information and conditions)</p>
        <p>3. Set analysis parameters and click "Run Analysis"</p>
        
        <h2>Analysis Steps</h2>
        <p><b>1. Data Loading:</b> The platform loads and filters low count genes.</p>
        <p><b>2. Normalization:</b> Counts are normalized using Counts Per Million (CPM).</p>
        <p><b>3. Differential Expression:</b> T-test is performed between selected conditions with FDR correction.</p>
        <p><b>4. Visualization:</b> PCA, heatmaps, and volcano plots are generated.</p>
        <p><b>5. Enrichment Analysis:</b> GO, KEGG, Reactome and Disease Ontology analysis can be run on significant genes.</p>
        
        <h2>Session Management</h2>
        <p>Save your analysis session (including parameters and results) to a JSON file for later use.</p>
        
        <h2>Exporting Results</h2>
        <p>Export results as PDF reports or Excel workbooks with multiple sheets.</p>
        
        <h2>Requirements</h2>
        <p>- Python 3.7+</p>
        <p>- Required packages: pandas, numpy, matplotlib, seaborn, scikit-learn, bioinfokit, gprofiler, reportlab, openpyxl</p>

        <h2>License</h2>
        <p>- Author: Victor Silveira Caricatte
        <p>- Universidade Federal De Minas Gerais- UFMG
        """
        
        self.help_text = QTextEdit()
        self.help_text.setReadOnly(True)
        self.help_text.setHtml(help_text)
        self.help_layout.addWidget(self.help_text)
        
        self.tabs.addTab(self.help_tab, "Help")
        
    def open_counts_file(self):
        try:
            file_path, _ = QFileDialog.getOpenFileName(self, "Open Counts File", "", "CSV Files (*.csv)")
            if file_path:
                self.counts_file = file_path
                self.counts_file_label.setText(f"Counts file: {os.path.basename(file_path)}")
                
                # Try to load the file to get conditions
                try:
                    df = pd.read_csv(file_path, index_col=0, nrows=1)  # Just read header
                    self.sample_names = df.columns.tolist()
                except pd.errors.EmptyDataError:
                    QMessageBox.warning(self, "Error", "The counts file is empty")
                except Exception as e:
                    QMessageBox.warning(self, "Error", f"Could not read counts file: {str(e)}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to open counts file: {str(e)}")
    
    def open_metadata_file(self):
        try:
            file_path, _ = QFileDialog.getOpenFileName(self, "Open Metadata File", "", "CSV Files (*.csv)")
            if file_path:
                self.metadata_file = file_path
                self.metadata_file_label.setText(f"Metadata file: {os.path.basename(file_path)}")
                
                # Load metadata to populate condition comboboxes
                try:
                    metadata = pd.read_csv(file_path)
                    
                    if 'condition' not in metadata.columns:
                        QMessageBox.warning(self, "Error", "Metadata file must contain a 'condition' column")
                        return
                    
                    if 'sample' not in metadata.columns:
                        QMessageBox.warning(self, "Error", "Metadata file must contain a 'sample' column")
                        return
                    
                    conditions = metadata['condition'].unique().tolist()
                    self.condition1_combo.clear()
                    self.condition2_combo.clear()
                    self.condition1_combo.addItems(conditions)
                    self.condition2_combo.addItems(conditions)
                except pd.errors.EmptyDataError:
                    QMessageBox.warning(self, "Error", "The metadata file is empty")
                except Exception as e:
                    QMessageBox.warning(self, "Error", f"Could not read metadata file: {str(e)}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to open metadata file: {str(e)}")
    
    def run_analysis(self):
        if not hasattr(self, 'counts_file') or not hasattr(self, 'metadata_file'):
            QMessageBox.warning(self, "Error", "Please select both counts and metadata files")
            return
            
        if self.condition1_combo.currentText() == "" or self.condition2_combo.currentText() == "":
            QMessageBox.warning(self, "Error", "Please select both conditions for comparison")
            return
            
        if self.condition1_combo.currentText() == self.condition2_combo.currentText():
            QMessageBox.warning(self, "Error", "Conditions must be different")
            return
            
        # Setup worker thread
        self.worker = RNAseqAnalysisWorker(
            counts_file=self.counts_file,
            metadata_file=self.metadata_file,
            conditions=[self.condition1_combo.currentText(), self.condition2_combo.currentText()],
            min_counts=self.min_counts_spin.value(),
            pval_cutoff=self.pval_cutoff_spin.value() / 100
        )
        
        self.worker.progress_updated.connect(self.update_progress)
        self.worker.analysis_completed.connect(self.analysis_finished)
        self.worker.error_occurred.connect(self.analysis_error)
        
        self.run_btn.setEnabled(False)
        self.progress_bar.setValue(0)
        self.worker.start()
    
    def update_progress(self, value):
        self.progress_bar.setValue(value)
    
    def analysis_finished(self, results):
        self.analysis_results = results
        self.run_btn.setEnabled(True)
        
        # Show results in tabs
        self.show_de_results()
        self.update_visualization()
        
        # Add results tabs if not already added
        if self.results_tab not in [self.tabs.widget(i) for i in range(self.tabs.count())]:
            self.tabs.addTab(self.results_tab, "Results")
        
        # Add sub-tabs to results tab
        self.results_tabs.addTab(self.de_tab, "Differential Expression")
        self.results_tabs.addTab(self.viz_tab, "Visualization")
        self.results_tabs.addTab(self.enrichment_tab, "Enrichment Analysis")
        
        QMessageBox.information(self, "Success", "Analysis completed successfully!")
    
    def analysis_error(self, error_msg):
        self.run_btn.setEnabled(True)
        error_dialog = QMessageBox(self)
        error_dialog.setIcon(QMessageBox.Icon.Critical)
        error_dialog.setWindowTitle("Analysis Error")
        error_dialog.setText("An error occurred during analysis")
        error_dialog.setDetailedText(error_msg)
        error_dialog.exec()
    
    def show_de_results(self):
        de_results = self.analysis_results['de_results']
        sig_genes = de_results[de_results['significant']]
        
        summary = f"""
        <h2>Differential Expression Results</h2>
        <p><b>Comparison:</b> {self.condition1_combo.currentText()} vs {self.condition2_combo.currentText()}</p>
        <p><b>Total genes:</b> {len(de_results)}</p>
        <p><b>Significant genes (FDR < {self.pval_cutoff_spin.value()/100:.3f}):</b> {len(sig_genes)}</p>
        <p><b>Top 10 significant genes:</b></p>
        """
        
        top_genes = sig_genes.sort_values('p-value').head(10)
        top_genes_html = top_genes[['mean_1', 'mean_2', 'log2fc', 'p-value', 'p-adj']].to_html(
            float_format=lambda x: f"{x:.4e}",
            formatters={
                'mean_1': '{:.2f}'.format,
                'mean_2': '{:.2f}'.format,
                'log2fc': '{:.2f}'.format
            }
        )
        
        self.de_text.setHtml(summary + top_genes_html)
    
    def update_visualization(self):
        if not self.analysis_results:
            return
            
        viz_type = self.viz_type_combo.currentText()
        fig = self.viz_canvas.figure
        fig.clf()
        ax = fig.add_subplot(111)
        
        if viz_type == "PCA Plot":
            self.plot_pca(ax)
        elif viz_type == "Heatmap":
            self.plot_heatmap(ax)
        elif viz_type == "Volcano Plot":
            self.plot_volcano(ax)
            
        # Enable interactive cursor
        mplcursors.cursor(hover=True)
        
        self.viz_canvas.draw()
    
    def plot_pca(self, ax):
        pca_data = self.analysis_results['pca']
        metadata = self.analysis_results['metadata']
        
        conditions = metadata['condition']
        unique_conditions = conditions.unique()
        colors = plt.cm.tab10(np.linspace(0, 1, len(unique_conditions)))
        condition_to_color = dict(zip(unique_conditions, colors))
        
        for condition in unique_conditions:
            idx = conditions == condition
            ax.scatter(
                pca_data['components'][idx, 0],
                pca_data['components'][idx, 1],
                color=condition_to_color[condition],
                label=condition,
                alpha=0.7
            )
            
        ax.set_xlabel(f"PC1 ({pca_data['explained_variance'][0]*100:.1f}%)")
        ax.set_ylabel(f"PC2 ({pca_data['explained_variance'][1]*100:.1f}%)")
        ax.set_title("PCA Plot")
        ax.legend()
        ax.grid(True)
    
    def plot_heatmap(self, ax):
        norm_counts = self.analysis_results['normalized_counts']
        metadata = self.analysis_results['metadata']
        
        # Get top 50 most variable genes
        top_genes = norm_counts.var(axis=1).sort_values(ascending=False).head(50).index
        heatmap_data = np.log2(norm_counts.loc[top_genes] + 1)
        
        # Cluster samples by condition
        conditions = metadata['condition']
        condition_order = conditions.sort_values().index
        heatmap_data = heatmap_data.iloc[:, condition_order]
        
        sns.heatmap(
            heatmap_data,
            ax=ax,
            cmap="viridis",
            yticklabels=True,
            xticklabels=heatmap_data.columns
        )
        ax.set_title("Top 50 Most Variable Genes (log2 CPM)")
        ax.set_xticklabels(ax.get_xticklabels(), rotation=45, ha='right')
    
    def plot_volcano(self, ax):
        de_results = self.analysis_results['de_results']
        pval_cutoff = self.pval_cutoff_spin.value() / 100
        
        # Calculate log2 fold change and -log10 p-value
        de_results['log2fc'] = np.log2(de_results['mean_1'] / de_results['mean_2'])
        de_results['neg_log10_pval'] = -np.log10(de_results['p-value'])
        
        # Color significant points
        colors = ['grey' if not sig else 'red' for sig in de_results['significant']]
        
        sc = ax.scatter(
            de_results['log2fc'],
            de_results['neg_log10_pval'],
            c=colors,
            alpha=0.5,
            s=10
        )
        
        ax.axhline(-np.log10(pval_cutoff), linestyle='--', color='black')
        ax.axvline(0, linestyle='--', color='black')
        
        ax.set_xlabel("log2 Fold Change")
        ax.set_ylabel("-log10 p-value")
        ax.set_title("Volcano Plot")
        
        # Add annotation for interactive points
        cursor = mplcursors.cursor(sc, hover=True)
        
        @cursor.connect("add")
        def on_add(sel):
            gene = de_results.index[sel.target.index]
            sel.annotation.set_text(
                f"{gene}\n"
                f"log2FC: {de_results.loc[gene, 'log2fc']:.2f}\n"
                f"p-value: {de_results.loc[gene, 'p-value']:.2e}\n"
                f"FDR: {de_results.loc[gene, 'p-adj']:.2e}"
            )
    
    def save_current_plot(self):
        if not hasattr(self, 'viz_canvas'):
            QMessageBox.warning(self, "Error", "No plot to save")
            return
            
        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            "Save Plot", 
            "", 
            "PNG Files (*.png);;PDF Files (*.pdf);;SVG Files (*.svg)"
        )
        
        if file_path:
            try:
                self.viz_canvas.figure.savefig(file_path, bbox_inches='tight', dpi=300)
                QMessageBox.information(self, "Success", f"Plot saved to {file_path}")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to save plot: {str(e)}")
    
    def run_enrichment_analysis(self):
        if not self.analysis_results or 'significant_genes' not in self.analysis_results:
            QMessageBox.warning(self, "Error", "Please run analysis first")
            return
            
        sig_genes = self.analysis_results['significant_genes']
        if not sig_genes:
            QMessageBox.information(self, "Info", "No significant genes found for enrichment analysis")
            return
            
        try:
            # Determine which sources to use
            sources = []
            if self.go_bp_check.isChecked():
                sources.append('GO:BP')
            if self.go_mf_check.isChecked():
                sources.append('GO:MF')
            if self.go_cc_check.isChecked():
                sources.append('GO:CC')
            if self.kegg_check.isChecked():
                sources.append('KEGG')
            if self.reactome_check.isChecked():
                sources.append('REAC')
            if self.do_check.isChecked():
                sources.append('DO')
            
            if not sources:
                QMessageBox.warning(self, "Error", "Please select at least one data source")
                return
                
            gp = GProfiler(return_dataframe=True)
            enrichment_results = gp.profile(
                organism='hsapiens',
                query=sig_genes,
                sources=sources
            )
            
            # Filter for significant results
            enrichment_results = enrichment_results[enrichment_results['p_value'] < 0.05]
            
            if enrichment_results.empty:
                self.enrichment_text.setHtml("<h2>Enrichment Results</h2><p>No significant terms found</p>")
                return
                
            # Format results
            summary = f"""
            <h2>Enrichment Results</h2>
            <p><b>Number of significant genes:</b> {len(sig_genes)}</p>
            <p><b>Number of significant terms:</b> {len(enrichment_results)}</p>
            <p><b>Top 10 terms:</b></p>
            """
            
            top_enrichment = enrichment_results.sort_values('p_value').head(10)
            top_enrichment_html = top_enrichment[['source', 'name', 'p_value', 'term_size', 'query_size']].to_html(
                float_format=lambda x: f"{x:.4e}",
                index=False
            )
            
            self.enrichment_text.setHtml(summary + top_enrichment_html)
            self.analysis_results['enrichment_results'] = enrichment_results
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Enrichment analysis failed: {str(e)}")
    
    def export_report(self):
        if not self.analysis_results:
            QMessageBox.warning(self, "Error", "No analysis results to export")
            return
            
        file_path, _ = QFileDialog.getSaveFileName(self, "Save Report", "", "PDF Files (*.pdf)")
        if not file_path:
            return
            
        try:
            # Create PDF document
            doc = SimpleDocTemplate(file_path, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []
            
            # Title
            title = Paragraph("RNA-Seq Analysis Report", styles['Title'])
            story.append(title)
            story.append(Spacer(1, 12))
            
            # Analysis info
            analysis_info = f"""
            <b>Analysis Parameters:</b><br/>
            Counts file: {os.path.basename(self.counts_file)}<br/>
            Metadata file: {os.path.basename(self.metadata_file)}<br/>
            Comparison: {self.condition1_combo.currentText()} vs {self.condition2_combo.currentText()}<br/>
            Minimum counts: {self.min_counts_spin.value()}<br/>
            P-value cutoff: {self.pval_cutoff_spin.value()/100:.3f}<br/>
            """
            story.append(Paragraph(analysis_info, styles['Normal']))
            story.append(Spacer(1, 12))
            
            # DE results
            de_title = Paragraph("Differential Expression Results", styles['Heading2'])
            story.append(de_title)
            
            de_results = self.analysis_results['de_results']
            sig_genes = de_results[de_results['significant']]
            
            de_summary = f"""
            <b>Total genes:</b> {len(de_results)}<br/>
            <b>Significant genes (FDR < {self.pval_cutoff_spin.value()/100:.3f}):</b> {len(sig_genes)}<br/>
            """
            story.append(Paragraph(de_summary, styles['Normal']))
            
            # Add top genes table
            top_genes = sig_genes.sort_values('p-value').head(10)
            table_data = [['Gene', 'Mean1', 'Mean2', 'log2FC', 'p-value', 'FDR']]
            for gene, row in top_genes.iterrows():
                table_data.append([
                    gene,
                    f"{row['mean_1']:.2f}",
                    f"{row['mean_2']:.2f}",
                    f"{row['log2fc']:.2f}",
                    f"{row['p-value']:.2e}",
                    f"{row['p-adj']:.2e}"
                ])
            
            de_table = Table(table_data)
            de_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(de_table)
            story.append(Spacer(1, 24))
            
            # Add visualizations
            viz_title = Paragraph("Visualizations", styles['Heading2'])
            story.append(viz_title)
            
            # Save PCA plot
            pca_fig = Figure(figsize=(8, 6))
            self.plot_pca(pca_fig.add_subplot(111))
            pca_path = os.path.join(self.temp_dir, "pca.png")
            pca_fig.savefig(pca_path, bbox_inches='tight', dpi=150)
            story.append(Image(pca_path, width=400, height=300))
            story.append(Spacer(1, 12))
            
            # Save volcano plot
            volcano_fig = Figure(figsize=(8, 6))
            self.plot_volcano(volcano_fig.add_subplot(111))
            volcano_path = os.path.join(self.temp_dir, "volcano.png")
            volcano_fig.savefig(volcano_path, bbox_inches='tight', dpi=150)
            story.append(Image(volcano_path, width=400, height=300))
            story.append(Spacer(1, 24))
            
            # Add enrichment results if available
            if 'enrichment_results' in self.analysis_results:
                enrichment_title = Paragraph("Enrichment Analysis Results", styles['Heading2'])
                story.append(enrichment_title)
                
                enrichment_results = self.analysis_results['enrichment_results']
                enrichment_summary = f"""
                <b>Number of significant terms:</b> {len(enrichment_results)}<br/>
                """
                story.append(Paragraph(enrichment_summary, styles['Normal']))
                
                # Add top enrichment terms table
                top_enrichment = enrichment_results.sort_values('p_value').head(10)
                table_data = [['Source', 'Term', 'Description', 'p-value', 'Term Size']]
                for _, row in top_enrichment.iterrows():
                    table_data.append([
                        row['source'],
                        row['native'],
                        row['name'],
                        f"{row['p_value']:.2e}",
                        str(row['term_size'])
                    ])
                
                enrichment_table = Table(table_data)
                enrichment_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(enrichment_table)
            
            # Build PDF
            doc.build(story)
            QMessageBox.information(self, "Success", f"Report saved to {file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export report: {str(e)}")
    
    def export_excel(self):
        if not self.analysis_results:
            QMessageBox.warning(self, "Error", "No analysis results to export")
            return
            
        file_path, _ = QFileDialog.getSaveFileName(self, "Save Excel Workbook", "", "Excel Files (*.xlsx)")
        if not file_path:
            return
            
        try:
            wb = Workbook()
            
            # Add DE results sheet
            de_sheet = wb.active
            de_sheet.title = "Differential Expression"
            
            de_results = self.analysis_results['de_results']
            for r in dataframe_to_rows(de_results.reset_index(), index=False, header=True):
                de_sheet.append(r)
            
            # Format header
            for cell in de_sheet[1]:
                cell.font = Font(bold=True)
            
            # Add normalized counts sheet
            norm_sheet = wb.create_sheet("Normalized Counts")
            norm_counts = self.analysis_results['normalized_counts']
            for r in dataframe_to_rows(norm_counts.reset_index(), index=False, header=True):
                norm_sheet.append(r)
            
            # Format header
            for cell in norm_sheet[1]:
                cell.font = Font(bold=True)
            
            # Add enrichment results if available
            if 'enrichment_results' in self.analysis_results:
                enrich_sheet = wb.create_sheet("Enrichment Results")
                enrich_results = self.analysis_results['enrichment_results']
                for r in dataframe_to_rows(enrich_results, index=False, header=True):
                    enrich_sheet.append(r)
                
                # Format header
                for cell in enrich_sheet[1]:
                    cell.font = Font(bold=True)
            
            # Save workbook
            wb.save(file_path)
            QMessageBox.information(self, "Success", f"Excel workbook saved to {file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export Excel workbook: {str(e)}")
    
    def save_session(self):
        if not hasattr(self, 'counts_file') or not hasattr(self, 'metadata_file'):
            QMessageBox.warning(self, "Error", "No analysis to save - please load files first")
            return
            
        dialog = SaveSessionDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            file_path = dialog.get_file_path()
            if not file_path:
                QMessageBox.warning(self, "Error", "Please select a file location")
                return
                
            try:
                session_data = {
                    'counts_file': self.counts_file,
                    'metadata_file': self.metadata_file,
                    'condition1': self.condition1_combo.currentText(),
                    'condition2': self.condition2_combo.currentText(),
                    'min_counts': self.min_counts_spin.value(),
                    'pval_cutoff': self.pval_cutoff_spin.value(),
                    'analysis_results': None
                }
                
                if dialog.include_raw_data() and self.analysis_results:
                    # Convert DataFrames to dictionaries for JSON serialization
                    results = {}
                    for key, value in self.analysis_results.items():
                        if isinstance(value, pd.DataFrame):
                            results[key] = value.to_dict()
                        elif isinstance(value, dict):
                            results[key] = value  # Will handle nested dicts like PCA results
                        else:
                            results[key] = str(value)  # Convert other types to string
                    
                    session_data['analysis_results'] = results
                
                with open(file_path, 'w') as f:
                    json.dump(session_data, f, indent=4)
                
                self.current_session_file = file_path
                QMessageBox.information(self, "Success", f"Session saved to {file_path}")
                
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to save session: {str(e)}")
    
    def load_session(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Load Session", "", "JSON Files (*.json)")
        if not file_path:
            return
            
        try:
            with open(file_path, 'r') as f:
                session_data = json.load(f)
            
            # Load basic session info
            self.counts_file = session_data['counts_file']
            self.counts_file_label.setText(f"Counts file: {os.path.basename(self.counts_file)}")
            
            self.metadata_file = session_data['metadata_file']
            self.metadata_file_label.setText(f"Metadata file: {os.path.basename(self.metadata_file)}")
            
            # Load metadata to populate condition comboboxes
            metadata = pd.read_csv(self.metadata_file)
            conditions = metadata['condition'].unique().tolist()
            self.condition1_combo.clear()
            self.condition2_combo.clear()
            self.condition1_combo.addItems(conditions)
            self.condition2_combo.addItems(conditions)
            
            # Set selected conditions
            self.condition1_combo.setCurrentText(session_data['condition1'])
            self.condition2_combo.setCurrentText(session_data['condition2'])
            
            # Set parameters
            self.min_counts_spin.setValue(session_data['min_counts'])
            self.pval_cutoff_spin.setValue(session_data['pval_cutoff'])
            
            # Load analysis results if available
            if 'analysis_results' in session_data and session_data['analysis_results']:
                results = {}
                for key, value in session_data['analysis_results'].items():
                    if isinstance(value, dict) and 'mean_1' in value:  # DE results
                        results[key] = pd.DataFrame.from_dict(value)
                    elif isinstance(value, dict) and 'components' in value:  # PCA results
                        results[key] = value  # Keep as dict
                    elif isinstance(value, dict) and 'index' in value:  # Other DataFrames
                        results[key] = pd.DataFrame.from_dict(value)
                    else:
                        results[key] = value
                
                self.analysis_results = results
                
                # Show results in tabs
                self.show_de_results()
                self.update_visualization()
                
                # Add results tabs if not already added
                if self.results_tab not in [self.tabs.widget(i) for i in range(self.tabs.count())]:
                    self.tabs.addTab(self.results_tab, "Results")
                
                # Add sub-tabs to results tab
                self.results_tabs.addTab(self.de_tab, "Differential Expression")
                self.results_tabs.addTab(self.viz_tab, "Visualization")
                self.results_tabs.addTab(self.enrichment_tab, "Enrichment Analysis")
            
            self.current_session_file = file_path
            QMessageBox.information(self, "Success", f"Session loaded from {file_path}")
            
        except FileNotFoundError:
            QMessageBox.critical(self, "Error", "The session file could not be found")
        except json.JSONDecodeError:
            QMessageBox.critical(self, "Error", "The session file is corrupted or invalid")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load session: {str(e)}")
    
    def closeEvent(self, event):
        # Clean up temporary directory
        if hasattr(self, 'temp_dir') and os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)
        
        # Stop any running worker thread
        if self.worker and self.worker.isRunning():
            self.worker.terminate()
        
        event.accept()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    platform = RNAseqAnalysisPlatform()
    platform.show()
    sys.exit(app.exec())
